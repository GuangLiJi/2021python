import openpyxl
workbook = openpyxl.load_workbook("C:\\Users\\jiche\\Desktop\\test.xlsx")
print(workbook.sheetnames)
for sheet in workbook :
    print(sheet.title)

newSheet = workbook.create_sheet("newSheetName")

print(workbook.sheetnames)

